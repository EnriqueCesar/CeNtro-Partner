#!/usr/bin/env python3
"""Valida Base_CeNtro Partner.xlsx por pestaña, encabezado y celda."""

from __future__ import annotations

import argparse
import hashlib
import json
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

MONTHS = ("ene", "feb", "mar", "abr", "may", "jun", "jul", "ago", "sep", "oct", "nov", "dic")
PERCENT_SHEETS = {"BB", "BT", "SS"}
REQUIRED_SHEETS = {
    "Directorio", "Instrucciones", "ADT_AA", "VMT_AA%", "V_ppto", "V_AT_AA%", "vCOGS",
    "SegundasCx", "NPS", "Conexion", "Desempeño", "Bebida", "SR% ", "Rotacion", "Bajas<90",
    "Estabilidad 12M", "Estabilidad 24M", "BB", "BT", "SS",
}


def normalized(value: Any) -> str:
    return str(value or "").strip().casefold()


def clean_ceco(value: Any) -> str:
    text = str(value or "").strip().removesuffix(".0")
    digits = "".join(character for character in text if character.isdigit())
    return digits.zfill(5) if digits else ""


def json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def audit_workbook(source: Path) -> dict[str, Any]:
    digest = hashlib.sha256(source.read_bytes()).hexdigest()
    workbook = load_workbook(source, data_only=False, read_only=True)
    missing_sheets = sorted(REQUIRED_SHEETS.difference(workbook.sheetnames))
    sheets: list[dict[str, Any]] = []
    totals = Counter(rows=0, cells=0, blanks=0, formulas=0, numbers=0, text=0, dates=0)

    for sheet in workbook.worksheets:
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        header_lookup = {normalized(header): index for index, header in enumerate(headers)}
        ceco_index = header_lookup.get("ceco")
        cecos: list[str] = []
        invalid_cecos: list[str] = []
        invalid_percentages: list[dict[str, Any]] = []
        counters = Counter(rows=max(sheet.max_row - 1, 0), cells=sheet.max_row * sheet.max_column)

        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            if ceco_index is not None:
                raw_ceco = row[ceco_index].value
                ceco = clean_ceco(raw_ceco)
                if ceco:
                    cecos.append(ceco)
                    if len(ceco) != 5:
                        invalid_cecos.append(str(raw_ceco))
            for column_number, cell in enumerate(row, start=1):
                value = cell.value
                if value in (None, ""):
                    counters["blanks"] += 1
                elif cell.data_type == "f":
                    counters["formulas"] += 1
                elif isinstance(value, bool):
                    counters["booleans"] += 1
                elif isinstance(value, (int, float)):
                    counters["numbers"] += 1
                elif isinstance(value, (datetime, date)):
                    counters["dates"] += 1
                else:
                    counters["text"] += 1

                header = normalized(headers[column_number - 1]) if column_number <= len(headers) else ""
                if sheet.title in PERCENT_SHEETS and header in MONTHS and value not in (None, "", "N/A", "NA"):
                    if not isinstance(value, (int, float)) or isinstance(value, bool) or not 0 <= float(value) <= 1:
                        invalid_percentages.append({"cell": cell.coordinate, "value": json_value(value)})

        duplicate_cecos = sorted(ceco for ceco, count in Counter(cecos).items() if count > 1)
        missing_headers = []
        if sheet.title == "Directorio":
            missing_headers = [header for header in ("CeCo", "Tienda", "Región", "DM") if normalized(header) not in header_lookup]
        elif sheet.title == "Instrucciones":
            expected = ("Pestaña", "Area", "Ponderacion", "Logica Selección Mes Multiple", "Logica YTD")
            missing_headers = [header for header in expected if normalized(header) not in header_lookup]
        elif sheet.title in REQUIRED_SHEETS:
            missing_headers = ["CeCo"] if "ceco" not in header_lookup else []

        totals.update(counters)
        sheets.append({
            "sheet": sheet.title,
            "rows": counters["rows"],
            "columns": sheet.max_column,
            "headers": [json_value(header) for header in headers],
            "missingHeaders": missing_headers,
            "validCeCos": len(set(cecos)),
            "duplicateCeCos": duplicate_cecos,
            "invalidCeCos": sorted(set(invalid_cecos)),
            "invalidPercentages": invalid_percentages,
            "types": {key: counters[key] for key in ("numbers", "text", "dates", "booleans", "formulas", "blanks")},
        })

    issues = sum(
        len(item["missingHeaders"]) + len(item["duplicateCeCos"]) + len(item["invalidCeCos"]) + len(item["invalidPercentages"])
        for item in sheets
    ) + len(missing_sheets)
    return {
        "schemaVersion": 1,
        "source": source.name,
        "sha256": digest,
        "workbook": {"sheetCount": len(workbook.sheetnames), "missingSheets": missing_sheets},
        "totals": dict(totals),
        "issueCount": issues,
        "sheets": sheets,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    report = audit_workbook(args.source)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(report, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(json.dumps({"issueCount": report["issueCount"], "output": str(args.output)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
